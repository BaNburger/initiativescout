#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import urldefrag, urljoin, urlsplit, urlunsplit

import pandas as pd
import requests
from lxml import html
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select

from initiative_tracker.db import init_db, session_scope
from initiative_tracker.models import DDEvidenceItem, DDTeamFact, DDTechFact, Initiative, InitiativePerson, InitiativeSource, Person, Score, Signal

DEFAULT_DB_URL = "sqlite:////Users/bastianburger/Repos/UnicornInitiative/data/initiatives.db"
DEFAULT_UNIVERSITIES = ("TUM", "LMU", "HM")
DEFAULT_OUTPUT = "/Users/bastianburger/Repos/UnicornInitiative/output/spreadsheet/student_initiatives_tum_lmu_hm_overview.xlsx"
DEFAULT_USER_AGENT = "UnicornInitiativeExport/1.0 (+https://unicorninitiative.local)"

UNIVERSITY_DOMAINS = {"tum.de", "lmu.de", "hm.edu"}

DOMAIN_CATEGORY_MAP = {
    "github": ("github.com",),
    "huggingface": ("huggingface.co",),
    "linkedin": ("linkedin.com",),
    "instagram": ("instagram.com",),
    "x_twitter": ("x.com", "twitter.com"),
    "facebook": ("facebook.com", "fb.me"),
    "youtube": ("youtube.com", "youtu.be"),
    "tiktok": ("tiktok.com",),
    "discord": ("discord.gg", "discord.com"),
    "researchgate": ("researchgate.net",),
    "openalex": ("openalex.org", "api.openalex.org"),
    "semantic_scholar": ("semanticscholar.org", "api.semanticscholar.org"),
}

OTHER_SOCIAL_DOMAINS = {
    "reddit.com",
    "meetup.com",
    "medium.com",
    "linktr.ee",
    "substack.com",
    "mastodon.social",
    "threads.net",
}

PROFILE_DISCOVERY_CATEGORIES = {
    "github",
    "huggingface",
    "linkedin",
    "instagram",
    "x_twitter",
    "facebook",
    "youtube",
    "tiktok",
    "discord",
    "researchgate",
    "other_social",
}

SCORE_DECIMALS = 4
TOKEN_STOPWORDS = {
    "and",
    "for",
    "the",
    "der",
    "die",
    "das",
    "und",
    "von",
    "zur",
    "zum",
    "mit",
    "munich",
    "muenchen",
    "student",
    "students",
    "initiative",
    "initiatives",
    "club",
    "team",
    "group",
    "association",
    "chapter",
    "faculty",
    "hochschule",
    "universitat",
    "university",
    "tum",
    "lmu",
    "hm",
}
GITHUB_GENERIC_SEGMENTS = {
    "",
    "about",
    "apps",
    "collections",
    "contact",
    "customer-stories",
    "enterprise",
    "events",
    "explore",
    "features",
    "issues",
    "join",
    "login",
    "marketplace",
    "new",
    "orgs",
    "organizations",
    "pricing",
    "readme",
    "search",
    "security",
    "settings",
    "site",
    "sponsors",
    "topics",
}


@dataclass
class ExportResult:
    output_path: Path
    rows: int
    website_crawls: int
    website_crawls_with_hits: int


def _parse_json_list(raw: str) -> list[str]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except Exception:  # noqa: BLE001
        return []
    if not isinstance(parsed, list):
        return []
    return [str(item) for item in parsed if isinstance(item, (str, int, float))]


def _domain(url: str) -> str:
    try:
        host = urlsplit(url).netloc.casefold().split(":", 1)[0]
    except Exception:  # noqa: BLE001
        return ""
    return host[4:] if host.startswith("www.") else host


def _domain_matches(host: str, target: str) -> bool:
    return host == target or host.endswith(f".{target}")


def _canonicalize_url(raw_url: str) -> str:
    if not raw_url:
        return ""
    raw = raw_url.strip()
    if not raw or not re.match(r"^https?://", raw, flags=re.IGNORECASE):
        return ""
    without_fragment, _ = urldefrag(raw)
    split = urlsplit(without_fragment)
    if not split.netloc:
        return ""
    path = split.path.rstrip("/")
    normalized_path = path if path else ""
    return urlunsplit((split.scheme.casefold(), split.netloc.casefold(), normalized_path, split.query, ""))


def _categorize_url(url: str) -> str:
    host = _domain(url)
    for category, domains in DOMAIN_CATEGORY_MAP.items():
        if any(_domain_matches(host, d) for d in domains):
            return category
    if any(_domain_matches(host, d) for d in OTHER_SOCIAL_DOMAINS):
        return "other_social"
    return "website"


def _is_university_domain(url: str) -> bool:
    host = _domain(url)
    return any(_domain_matches(host, domain) for domain in UNIVERSITY_DOMAINS)


def _is_directory_page(url: str) -> bool:
    lower = url.casefold()
    return (
        "student-clubs-galerie" in lower
        or "studentische-initiativen" in lower
        or "aktivitaeten" in lower
    )


def _extract_evidence_urls(raw_json: str) -> list[str]:
    out: list[str] = []
    if not raw_json:
        return out
    try:
        parsed = json.loads(raw_json)
    except Exception:  # noqa: BLE001
        return out
    if not isinstance(parsed, list):
        return out
    for item in parsed:
        if not isinstance(item, dict):
            continue
        source_url = _canonicalize_url(str(item.get("source_url") or ""))
        if source_url:
            out.append(source_url)
    return out


def _join(items: Iterable[str]) -> str:
    values = [item for item in items if item]
    return "; ".join(sorted(set(values)))


def _tokenize(text: str) -> set[str]:
    tokens: set[str] = set()
    for token in re.split(r"[^a-z0-9]+", text.casefold()):
        if len(token) < 3:
            continue
        if token in TOKEN_STOPWORDS:
            continue
        tokens.add(token)
    return tokens


def _initiative_relevance_tokens(name: str, urls: Iterable[str]) -> set[str]:
    tokens = _tokenize(name)
    for url in urls:
        host = _domain(url)
        tokens.update(_tokenize(host))
    return tokens


def _is_plausible_profile_url(url: str, category: str, relevance_tokens: set[str]) -> bool:
    if category not in {"github", "huggingface", "openalex", "semantic_scholar"}:
        return True

    split = urlsplit(url)
    path = split.path.casefold().strip("/")
    if not path:
        return False

    if category == "github":
        first = path.split("/", 1)[0]
        if first in GITHUB_GENERIC_SEGMENTS:
            return False

    if not relevance_tokens:
        return True
    return any(token in path for token in relevance_tokens)


def _top_signals(rows: list[Signal], signal_type: str, limit: int = 8) -> list[str]:
    bucket: dict[str, float] = defaultdict(float)
    for row in rows:
        if row.signal_type != signal_type:
            continue
        bucket[row.signal_key] += float(row.value)
    ordered = sorted(bucket.items(), key=lambda kv: kv[1], reverse=True)
    return [key for key, _ in ordered[:limit]]


def _latest_by(rows: list, key_attr: str) -> dict[int, object]:
    out: dict[int, object] = {}
    ordered = sorted(rows, key=lambda row: getattr(row, key_attr) or datetime(1970, 1, 1, tzinfo=UTC), reverse=True)
    for row in ordered:
        out.setdefault(int(row.initiative_id), row)
    return out


def _discover_profile_links(seed_url: str, *, timeout: float, user_agent: str) -> set[str]:
    headers = {"User-Agent": user_agent}
    try:
        response = requests.get(seed_url, headers=headers, timeout=timeout)
    except Exception:  # noqa: BLE001
        return set()
    if response.status_code >= 400 or not response.text:
        return set()
    try:
        tree = html.fromstring(response.text)
    except Exception:  # noqa: BLE001
        return set()

    found: set[str] = set()
    for href in tree.xpath("//a[@href]/@href"):
        absolute = _canonicalize_url(urljoin(seed_url, str(href)))
        if not absolute:
            continue
        if _categorize_url(absolute) in PROFILE_DISCOVERY_CATEGORIES:
            found.add(absolute)
    return found


def _build_output_rows(
    *,
    initiatives: list[Initiative],
    source_map: dict[int, list[InitiativeSource]],
    signal_map: dict[int, list[Signal]],
    links_map: dict[int, list[InitiativePerson]],
    people_by_id: dict[int, Person],
    latest_scores: dict[int, Score],
    latest_dd_team: dict[int, DDTeamFact],
    latest_dd_tech: dict[int, DDTechFact],
    evidence_map: dict[int, list[DDEvidenceItem]],
    crawl_timeout: float,
    crawl_workers: int,
    crawl_websites: bool,
) -> tuple[list[dict[str, object]], int, int]:
    initial_urls: dict[int, set[str]] = {}
    crawl_seeds: dict[int, list[str]] = {}

    for initiative in initiatives:
        urls: set[str] = set()

        primary = _canonicalize_url(initiative.primary_url)
        if primary:
            urls.add(primary)

        for source in source_map.get(initiative.id, []):
            source_url = _canonicalize_url(source.source_url)
            external_url = _canonicalize_url(source.external_url)
            if source_url:
                urls.add(source_url)
            if external_url:
                urls.add(external_url)

        for link in links_map.get(initiative.id, []):
            person = people_by_id.get(link.person_id)
            if not person:
                continue
            for source_url in _parse_json_list(person.source_urls_json):
                normalized = _canonicalize_url(source_url)
                if normalized:
                    urls.add(normalized)
            for channel in _parse_json_list(person.contact_channels_json):
                normalized = _canonicalize_url(channel)
                if normalized:
                    urls.add(normalized)

        tech = latest_dd_tech.get(initiative.id)
        if tech:
            if tech.source_url:
                normalized = _canonicalize_url(tech.source_url)
                if normalized:
                    urls.add(normalized)
            if tech.github_org:
                github_url = (
                    f"https://github.com/{tech.github_org.strip()}/{tech.github_repo.strip()}"
                    if tech.github_repo.strip()
                    else f"https://github.com/{tech.github_org.strip()}"
                )
                normalized = _canonicalize_url(github_url)
                if normalized:
                    urls.add(normalized)
            for evidence_url in _extract_evidence_urls(tech.evidence_json):
                urls.add(evidence_url)

        team = latest_dd_team.get(initiative.id)
        if team:
            if team.source_url:
                normalized = _canonicalize_url(team.source_url)
                if normalized:
                    urls.add(normalized)
            for evidence_url in _extract_evidence_urls(team.evidence_json):
                urls.add(evidence_url)

        for evidence in evidence_map.get(initiative.id, []):
            normalized = _canonicalize_url(evidence.source_url)
            if normalized:
                urls.add(normalized)

        initial_urls[initiative.id] = urls

        website_candidates = [
            url
            for url in sorted(urls)
            if _categorize_url(url) == "website"
            and not _is_directory_page(url)
            and not url.casefold().endswith((".pdf", ".jpg", ".jpeg", ".png", ".svg"))
        ]
        non_university = [url for url in website_candidates if not _is_university_domain(url)]
        chosen = non_university[:1] or website_candidates[:1]
        crawl_seeds[initiative.id] = chosen

    crawls = 0
    crawls_with_hits = 0
    discovered: dict[int, set[str]] = defaultdict(set)

    if crawl_websites:
        seed_to_ids: dict[str, set[int]] = defaultdict(set)
        for initiative_id, seeds in crawl_seeds.items():
            for seed in seeds:
                seed_to_ids[seed].add(initiative_id)

        if seed_to_ids:
            with ThreadPoolExecutor(max_workers=max(1, crawl_workers)) as executor:
                futures = {
                    executor.submit(
                        _discover_profile_links,
                        seed,
                        timeout=crawl_timeout,
                        user_agent=DEFAULT_USER_AGENT,
                    ): seed
                    for seed in seed_to_ids
                }
                for future in as_completed(futures):
                    seed = futures[future]
                    crawls += 1
                    links = future.result()
                    if links:
                        crawls_with_hits += 1
                    for initiative_id in seed_to_ids[seed]:
                        discovered[initiative_id].update(links)

    rows: list[dict[str, object]] = []
    for initiative in initiatives:
        all_urls = set(initial_urls.get(initiative.id, set()))
        all_urls.update(discovered.get(initiative.id, set()))

        website_relevance_urls = [url for url in all_urls if _categorize_url(url) == "website"]
        relevance_tokens = _initiative_relevance_tokens(initiative.canonical_name, website_relevance_urls)
        categories: dict[str, list[str]] = defaultdict(list)
        for url in all_urls:
            category = _categorize_url(url)
            if not _is_plausible_profile_url(url, category, relevance_tokens):
                continue
            categories[category].append(url)

        signals = signal_map.get(initiative.id, [])
        tech_domains = _top_signals(signals, "technology_domain", limit=10)
        market_domains = _top_signals(signals, "market_domain", limit=10)
        categories_raw = _parse_json_list(initiative.categories_json)

        link_rows = links_map.get(initiative.id, [])
        people_ids = sorted({link.person_id for link in link_rows})
        people_names = [people_by_id[pid].canonical_name for pid in people_ids if pid in people_by_id]
        roles = [link.role.strip() for link in link_rows if link.role and link.role.strip()]

        score = latest_scores.get(initiative.id)
        dd_team = latest_dd_team.get(initiative.id)
        dd_tech = latest_dd_tech.get(initiative.id)

        evidence_counter: Counter[str] = Counter()
        for item in evidence_map.get(initiative.id, []):
            evidence_counter[item.source_type] += 1

        directory_sources = [
            src.source_url
            for src in source_map.get(initiative.id, [])
            if src.source_type == "directory_scrape" and src.source_url
        ]
        website_sources = [
            src.external_url
            for src in source_map.get(initiative.id, [])
            if src.source_type in {"directory_scrape", "website_enrichment"} and src.external_url
        ]

        profile_coverage = sum(
            int(bool(categories.get(key)))
            for key in [
                "github",
                "huggingface",
                "linkedin",
                "instagram",
                "x_twitter",
                "facebook",
                "youtube",
                "tiktok",
                "discord",
                "researchgate",
            ]
        )

        row = {
            "university": initiative.university,
            "initiative_id": int(initiative.id),
            "initiative_name": initiative.canonical_name,
            "primary_url": _canonicalize_url(initiative.primary_url),
            "directory_source_urls": _join(directory_sources),
            "website_urls": _join(categories.get("website", [])),
            "github_urls": _join(categories.get("github", [])),
            "huggingface_urls": _join(categories.get("huggingface", [])),
            "linkedin_urls": _join(categories.get("linkedin", [])),
            "instagram_urls": _join(categories.get("instagram", [])),
            "x_twitter_urls": _join(categories.get("x_twitter", [])),
            "facebook_urls": _join(categories.get("facebook", [])),
            "youtube_urls": _join(categories.get("youtube", [])),
            "tiktok_urls": _join(categories.get("tiktok", [])),
            "discord_urls": _join(categories.get("discord", [])),
            "researchgate_urls": _join(categories.get("researchgate", [])),
            "openalex_urls": _join(categories.get("openalex", [])),
            "semantic_scholar_urls": _join(categories.get("semantic_scholar", [])),
            "other_social_urls": _join(categories.get("other_social", [])),
            "profile_coverage_score": profile_coverage,
            "technology_domains": _join(tech_domains),
            "market_domains": _join(market_domains),
            "categories": _join(categories_raw),
            "activity_summary": (initiative.description_summary_en or initiative.description_raw or "").strip(),
            "member_count_detected": len(people_ids),
            "member_examples": _join(people_names[:8]),
            "member_roles_detected": _join(roles[:12]),
            "dd_key_roles": _join(_parse_json_list(dd_team.key_roles_json) if dd_team else []),
            "dd_references_count": int(dd_team.references_count) if dd_team else 0,
            "dd_is_investable": bool(dd_team.is_investable) if dd_team else False,
            "github_repo_count": int(dd_tech.repo_count) if dd_tech else 0,
            "github_contributors": int(dd_tech.contributor_count) if dd_tech else 0,
            "github_commits_90d": round(float(dd_tech.commit_velocity_90d), SCORE_DECIMALS) if dd_tech else 0.0,
            "github_ci_present": bool(dd_tech.ci_present) if dd_tech else False,
            "huggingface_model_hits": int(evidence_counter.get("huggingface", 0)),
            "openalex_hits": int(evidence_counter.get("openalex", 0)),
            "semantic_scholar_hits": int(evidence_counter.get("semantic_scholar", 0)),
            "linkedin_hits": int(evidence_counter.get("linkedin_safe", 0)),
            "researchgate_hits": int(evidence_counter.get("researchgate_safe", 0)),
            "outreach_now_score": round(float(score.outreach_now_score), SCORE_DECIMALS) if score else None,
            "venture_upside_score": round(float(score.venture_upside_score), SCORE_DECIMALS) if score else None,
            "last_seen_at": initiative.last_seen_at.isoformat() if initiative.last_seen_at else "",
            "all_source_urls": _join(all_urls),
            "known_url_count": len(all_urls),
            "seed_website_urls": _join(website_sources),
        }
        rows.append(row)

    rows.sort(key=lambda row: (str(row["university"]), str(row["initiative_name"]).casefold()))
    return rows, crawls, crawls_with_hits


def _style_workbook(path: Path, sheet_name: str) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in col_cells]
        width = min(90, max(12, max(len(value) for value in values) + 2))
        worksheet.column_dimensions[col_cells[0].column_letter].width = width

    workbook.save(path)


def _build_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for university, group in df.groupby("university", dropna=False):
        rows.append(
            {
                "university": university,
                "initiative_count": int(len(group)),
                "with_website": int(group["website_urls"].astype(bool).sum()),
                "with_github": int(group["github_urls"].astype(bool).sum()),
                "with_huggingface": int(group["huggingface_urls"].astype(bool).sum()),
                "with_linkedin": int(group["linkedin_urls"].astype(bool).sum()),
                "with_other_social": int(group["other_social_urls"].astype(bool).sum()),
                "with_members_detected": int((group["member_count_detected"] > 0).sum()),
                "avg_profile_coverage": round(float(group["profile_coverage_score"].mean()), 3),
                "avg_known_url_count": round(float(group["known_url_count"].mean()), 2),
            }
        )
    return pd.DataFrame(rows).sort_values("university")


def export_university_overview(
    *,
    db_url: str,
    output_path: Path,
    universities: tuple[str, ...],
    crawl_websites: bool,
    crawl_timeout: float,
    crawl_workers: int,
) -> ExportResult:
    init_db(db_url)
    with session_scope(db_url) as session:
        initiatives = (
            session.execute(select(Initiative).where(Initiative.university.in_(universities)))
            .scalars()
            .all()
        )
        initiative_ids = [initiative.id for initiative in initiatives]
        if not initiative_ids:
            raise RuntimeError(f"No initiatives found for universities: {', '.join(universities)}")

        sources = (
            session.execute(select(InitiativeSource).where(InitiativeSource.initiative_id.in_(initiative_ids)))
            .scalars()
            .all()
        )
        signals = (
            session.execute(select(Signal).where(Signal.initiative_id.in_(initiative_ids)))
            .scalars()
            .all()
        )
        links = (
            session.execute(select(InitiativePerson).where(InitiativePerson.initiative_id.in_(initiative_ids)))
            .scalars()
            .all()
        )
        scores = (
            session.execute(select(Score).where(Score.initiative_id.in_(initiative_ids)))
            .scalars()
            .all()
        )
        dd_team = (
            session.execute(select(DDTeamFact).where(DDTeamFact.initiative_id.in_(initiative_ids)))
            .scalars()
            .all()
        )
        dd_tech = (
            session.execute(select(DDTechFact).where(DDTechFact.initiative_id.in_(initiative_ids)))
            .scalars()
            .all()
        )
        dd_evidence = (
            session.execute(select(DDEvidenceItem).where(DDEvidenceItem.initiative_id.in_(initiative_ids)))
            .scalars()
            .all()
        )

        person_ids = sorted({row.person_id for row in links})
        people_rows = (
            session.execute(select(Person).where(Person.id.in_(person_ids))).scalars().all()
            if person_ids
            else []
        )

    source_map: dict[int, list[InitiativeSource]] = defaultdict(list)
    for row in sources:
        source_map[row.initiative_id].append(row)

    signal_map: dict[int, list[Signal]] = defaultdict(list)
    for row in signals:
        signal_map[row.initiative_id].append(row)

    links_map: dict[int, list[InitiativePerson]] = defaultdict(list)
    for row in links:
        links_map[row.initiative_id].append(row)

    people_by_id = {person.id: person for person in people_rows}
    latest_scores = _latest_by(scores, "scored_at")
    latest_dd_team = _latest_by(dd_team, "updated_at")
    latest_dd_tech = _latest_by(dd_tech, "updated_at")

    evidence_map: dict[int, list[DDEvidenceItem]] = defaultdict(list)
    for row in dd_evidence:
        evidence_map[row.initiative_id].append(row)

    rows, crawl_count, crawl_hits = _build_output_rows(
        initiatives=initiatives,
        source_map=source_map,
        signal_map=signal_map,
        links_map=links_map,
        people_by_id=people_by_id,
        latest_scores=latest_scores,
        latest_dd_team=latest_dd_team,
        latest_dd_tech=latest_dd_tech,
        evidence_map=evidence_map,
        crawl_timeout=crawl_timeout,
        crawl_workers=crawl_workers,
        crawl_websites=crawl_websites,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    frame = pd.DataFrame(rows)
    summary = _build_summary(frame)
    generated_meta = pd.DataFrame(
        [
            {"metric": "generated_at_utc", "value": datetime.now(tz=UTC).isoformat()},
            {"metric": "db_url", "value": db_url},
            {"metric": "universities", "value": ", ".join(universities)},
            {"metric": "rows", "value": len(frame)},
            {"metric": "website_crawls", "value": crawl_count},
            {"metric": "website_crawls_with_hits", "value": crawl_hits},
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="Initiatives")
        summary.to_excel(writer, index=False, sheet_name="Summary")
        generated_meta.to_excel(writer, index=False, sheet_name="Meta")

    _style_workbook(output_path, "Initiatives")
    return ExportResult(
        output_path=output_path,
        rows=len(frame),
        website_crawls=crawl_count,
        website_crawls_with_hits=crawl_hits,
    )


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export TUM/LMU/HM student initiatives with website + GitHub/HuggingFace/social profile overview to Excel."
    )
    parser.add_argument("--db-url", default=DEFAULT_DB_URL, help="SQLAlchemy DB URL")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output .xlsx path")
    parser.add_argument(
        "--universities",
        default="TUM,LMU,HM",
        help="Comma-separated university codes to include",
    )
    parser.add_argument(
        "--skip-website-crawl",
        action="store_true",
        help="Disable outbound profile link discovery from initiative websites",
    )
    parser.add_argument("--crawl-timeout", type=float, default=6.0, help="HTTP timeout in seconds for website link discovery")
    parser.add_argument("--crawl-workers", type=int, default=12, help="Thread count for website link discovery")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    universities = tuple(sorted({token.strip().upper() for token in args.universities.split(",") if token.strip()}))
    if not universities:
        raise SystemExit("No universities provided")

    result = export_university_overview(
        db_url=str(args.db_url),
        output_path=Path(args.output).expanduser().resolve(),
        universities=universities,
        crawl_websites=not bool(args.skip_website_crawl),
        crawl_timeout=float(args.crawl_timeout),
        crawl_workers=int(args.crawl_workers),
    )
    print(
        json.dumps(
            {
                "status": "ok",
                "output": str(result.output_path),
                "rows": result.rows,
                "website_crawls": result.website_crawls,
                "website_crawls_with_hits": result.website_crawls_with_hits,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
