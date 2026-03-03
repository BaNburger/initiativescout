"""XLSX export: write initiative data (with scores and enrichments) to a workbook."""
from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from scout.models import Enrichment, Initiative, OutreachScore

# ---------------------------------------------------------------------------
# Column definitions (order matters — determines sheet layout)
# ---------------------------------------------------------------------------

# Core profile columns
_PROFILE_COLS: list[tuple[str, str, int]] = [
    # (header, attribute_or_key, width)
    ("ID", "id", 6),
    ("Name", "name", 30),
    ("University", "uni", 10),
    ("Faculty", "faculty", 18),
    ("Sector", "sector", 16),
    ("Mode", "mode", 10),
    ("Description", "description", 40),
    ("Website", "website", 30),
    ("Email", "email", 25),
    ("LinkedIn", "linkedin", 30),
    ("GitHub Org", "github_org", 20),
    ("Team Page", "team_page", 30),
    ("Team Size", "team_size", 10),
]

# Score columns (from latest OutreachScore)
_SCORE_COLS: list[tuple[str, str, int]] = [
    ("Verdict", "verdict", 14),
    ("Score", "score", 8),
    ("Classification", "classification", 16),
    ("Grade Team", "grade_team", 10),
    ("Grade Tech", "grade_tech", 10),
    ("Grade Opp", "grade_opportunity", 10),
    ("Reasoning", "reasoning", 50),
    ("Contact", "contact_who", 25),
    ("Channel", "contact_channel", 12),
    ("Engagement Hook", "engagement_hook", 40),
]

# Extra profile fields
_EXTRA_COLS: list[tuple[str, str, int]] = [
    ("Relevance", "relevance", 12),
    ("Key Repos", "key_repos", 30),
    ("Sponsors", "sponsors", 25),
    ("Competitions", "competitions", 25),
    ("Tech Domains", "technology_domains", 25),
    ("Market Domains", "market_domains", 25),
    ("Categories", "categories", 20),
    ("Member Count", "member_count", 12),
]

# Enrichment summary column (concatenated)
_ENRICHMENT_COL = ("Enrichment Summary", 50)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="e0e0e0", size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")

_VERDICT_FILLS = {
    "reach_out_now": PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid"),
    "reach_out_soon": PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid"),
    "monitor": PatternFill(start_color="e0e7ff", end_color="e0e7ff", fill_type="solid"),
    "skip": PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid"),
}


# ---------------------------------------------------------------------------
# Export logic
# ---------------------------------------------------------------------------


def _latest_scores(session: Session) -> dict[int, OutreachScore]:
    """Load the latest initiative-level score per initiative."""
    subq = (
        select(
            OutreachScore.initiative_id,
            func.max(OutreachScore.scored_at).label("max_scored"),
        )
        .where(OutreachScore.project_id.is_(None))
        .group_by(OutreachScore.initiative_id)
        .subquery()
    )
    rows = session.execute(
        select(OutreachScore)
        .join(subq, (OutreachScore.initiative_id == subq.c.initiative_id)
              & (OutreachScore.scored_at == subq.c.max_scored))
        .where(OutreachScore.project_id.is_(None))
    ).scalars().all()
    return {s.initiative_id: s for s in rows}


def _enrichment_summaries(session: Session) -> dict[int, str]:
    """Concatenate enrichment summaries per initiative."""
    rows = session.execute(
        select(Enrichment.initiative_id, Enrichment.source_type, Enrichment.summary)
        .order_by(Enrichment.initiative_id, Enrichment.source_type)
    ).all()
    result: dict[int, list[str]] = {}
    for init_id, source, summary in rows:
        if summary:
            result.setdefault(init_id, []).append(f"[{source}] {summary}")
    return {k: "\n\n".join(v) for k, v in result.items()}


def export_xlsx(
    session: Session,
    *,
    verdict: str | None = None,
    uni: str | None = None,
    include_enrichments: bool = True,
    include_scores: bool = True,
    include_extras: bool = False,
) -> BytesIO:
    """Export initiatives to an XLSX workbook returned as a BytesIO buffer.

    Args:
        session: DB session.
        verdict: Comma-separated verdict filter (e.g. "reach_out_now,reach_out_soon").
        uni: Comma-separated uni filter (e.g. "TUM,LMU").
        include_enrichments: Include enrichment summary column.
        include_scores: Include score columns (verdict, grades, reasoning, etc.).
        include_extras: Include extra profile fields (domains, member count, etc.).
    """
    # Pre-load related data in bulk (before filtering, so verdict filter can reuse)
    score_map = _latest_scores(session) if (include_scores or verdict) else {}
    enrich_map = _enrichment_summaries(session) if include_enrichments else {}

    # Build query
    query = select(Initiative).order_by(Initiative.uni, Initiative.name)
    if verdict:
        wanted = {v.strip().lower() for v in verdict.split(",")}
        verdict_ids = [sid for sid, s in score_map.items() if s.verdict in wanted]
        if "unscored" in wanted:
            scored_ids = set(score_map.keys())
            all_ids = {r[0] for r in session.execute(select(Initiative.id)).all()}
            verdict_ids.extend(all_ids - scored_ids)
        query = query.where(Initiative.id.in_(verdict_ids))
    if uni:
        unis = {u.strip().upper() for u in uni.split(",")}
        query = query.where(func.upper(Initiative.uni).in_(unis))

    initiatives = session.execute(query).scalars().all()

    # Build columns list
    columns: list[tuple[str, str, int]] = list(_PROFILE_COLS)
    if include_scores:
        columns.extend(_SCORE_COLS)
    if include_extras:
        columns.extend(_EXTRA_COLS)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Initiatives"

    # Header row
    headers = [c[0] for c in columns]
    if include_enrichments:
        headers.append(_ENRICHMENT_COL[0])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    # Set column widths
    widths = [c[2] for c in columns]
    if include_enrichments:
        widths.append(_ENRICHMENT_COL[1])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for init in initiatives:
        score = score_map.get(init.id)
        row: list[Any] = []
        for _, attr, _ in columns:
            if attr in ("verdict", "score", "classification", "reasoning",
                        "contact_who", "contact_channel", "engagement_hook",
                        "grade_team", "grade_tech", "grade_opportunity"):
                row.append(getattr(score, attr, None) if score else None)
            else:
                val = getattr(init, attr, "")
                # Convert booleans for Excel
                if isinstance(val, bool):
                    val = "Yes" if val else "No"
                row.append(val)
        if include_enrichments:
            row.append(enrich_map.get(init.id, ""))
        ws.append(row)

        # Style verdict cell
        if include_scores and score and score.verdict in _VERDICT_FILLS:
            verdict_col = next(
                (i for i, (_, attr, _) in enumerate(columns) if attr == "verdict"), None
            )
            if verdict_col is not None:
                ws.cell(row=ws.max_row, column=verdict_col + 1).fill = _VERDICT_FILLS[score.verdict]

    # Wrap text for long columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = _WRAP

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
