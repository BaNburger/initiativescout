from __future__ import annotations

import json
import logging
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from scout.models import ImportResult, Initiative

log = logging.getLogger(__name__)


def _s(value: object) -> str:
    """Safely coerce cell value to stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def _col(row: tuple, idx: int) -> object:
    """Safely get a column value from a row tuple."""
    return row[idx] if idx < len(row) else None


def _parse_spin_off_sheet(ws) -> list[dict]:
    """Parse the 'Spin-Off Targets' sheet.

    Current layout (1 header row):
    [0]=Name [1]=University [2]=Sector [3]=Description [4]=Website [5]=Email
    [6]=LinkedIn [7]=Instagram [8]=X/Twitter [9]=Discord [10]=Facebook
    [11]=YouTube [12]=GitHub [13]=TikTok [14]=Slack [15]=Activity Mode
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out: list[dict] = []
    for row in rows:
        if not row or not _col(row, 0):  # col 0 = Name
            continue
        out.append({
            "name": _s(_col(row, 0)),
            "uni": _s(_col(row, 1)),
            "sector": _s(_col(row, 2)),
            "mode": _s(_col(row, 15)),
            "description": _s(_col(row, 3)),
            "website": _s(_col(row, 4)),
            "email": _s(_col(row, 5)),
            "linkedin": _s(_col(row, 6)),
            "github_org": _s(_col(row, 12)),
            "team_page": "",
            "team_size": "",
            "key_repos": "",
            "sponsors": "",
            "competitions": "",
            "relevance": "",
            "sheet_source": "spin_off_targets",
            "extra_links_json": json.dumps({
                k: v for k, v in {
                    "instagram": _s(_col(row, 7)),
                    "x_twitter": _s(_col(row, 8)),
                    "discord": _s(_col(row, 9)),
                    "facebook": _s(_col(row, 10)),
                    "youtube": _s(_col(row, 11)),
                    "tiktok": _s(_col(row, 13)),
                    "slack": _s(_col(row, 14)),
                }.items() if v
            }),
        })
    return out


def _parse_all_initiatives_sheet(ws) -> list[dict]:
    """Parse the 'All Initiatives' sheet.

    Current layout (1 header row):
    [0]=Name [1]=Spinoff Relevance [2]=University [3]=Sector [4]=Description
    [5]=Website [6]=Email [7]=LinkedIn [8]=Instagram [9]=X/Twitter [10]=Discord
    [11]=Facebook [12]=YouTube [13]=GitHub [14]=TikTok [15]=HuggingFace
    [16]=Luma [17]=Linktree [18]=Slack [19]=Activity Mode
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out: list[dict] = []
    for row in rows:
        if not row or not _col(row, 0):  # col 0 = Name
            continue
        out.append({
            "name": _s(_col(row, 0)),
            "uni": _s(_col(row, 2)),
            "sector": _s(_col(row, 3)),
            "mode": _s(_col(row, 19)),
            "description": _s(_col(row, 4)),
            "website": _s(_col(row, 5)),
            "email": _s(_col(row, 6)),
            "linkedin": _s(_col(row, 7)),
            "github_org": _s(_col(row, 13)),
            "team_page": "",
            "team_size": "",
            "key_repos": "",
            "sponsors": "",
            "competitions": "",
            "relevance": _s(_col(row, 1)),
            "sheet_source": "all_initiatives",
            "extra_links_json": json.dumps({
                k: v for k, v in {
                    "instagram": _s(_col(row, 8)),
                    "x_twitter": _s(_col(row, 9)),
                    "discord": _s(_col(row, 10)),
                    "facebook": _s(_col(row, 11)),
                    "youtube": _s(_col(row, 12)),
                    "tiktok": _s(_col(row, 14)),
                    "huggingface": _s(_col(row, 15)),
                    "luma": _s(_col(row, 16)),
                    "linktree": _s(_col(row, 17)),
                    "slack": _s(_col(row, 18)),
                }.items() if v
            }),
        })
    return out


def _i(value: object) -> int:
    """Safely coerce cell value to int."""
    if value is None:
        return 0
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0


def _f(value: object) -> float | None:
    """Safely coerce cell value to float, None if missing."""
    if value is None:
        return None
    try:
        v = float(value)
        return v if v != 0.0 else None
    except (ValueError, TypeError):
        return None


def _b(value: object) -> bool:
    """Safely coerce cell value to bool."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes")


def _parse_overview_sheet(ws) -> list[dict]:
    """Parse the 'Initiatives' sheet from the overview spreadsheet (1 header row)."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out: list[dict] = []
    for row in rows:
        if not row or not _col(row, 2):  # col 2 = initiative_name
            continue
        # Collect URL fields into extra_links
        url_links = {}
        for key, idx in (
            ("primary_url", 3), ("directory_source_urls", 4), ("website_urls", 5),
            ("github_urls", 6), ("huggingface_urls", 7), ("linkedin_urls", 8),
            ("instagram_urls", 9), ("x_twitter_urls", 10), ("facebook_urls", 11),
            ("youtube_urls", 12), ("tiktok_urls", 13), ("discord_urls", 14),
            ("researchgate_urls", 15), ("openalex_urls", 16),
            ("semantic_scholar_urls", 17), ("other_social_urls", 18),
        ):
            val = _s(_col(row, idx))
            if val:
                url_links[key] = val

        out.append({
            "name": _s(_col(row, 2)),
            "uni": _s(_col(row, 0)),
            "sheet_source": "overview",
            "extra_links_json": json.dumps(url_links),
            # Classification
            "technology_domains": _s(_col(row, 20)),
            "market_domains": _s(_col(row, 21)),
            "categories": _s(_col(row, 22)),
            # Activity summary → description fallback
            "description": _s(_col(row, 23)),
            # Team signals
            "member_count": _i(_col(row, 24)),
            "member_examples": _s(_col(row, 25)),
            "member_roles": _s(_col(row, 26)),
            # Due diligence
            "dd_key_roles": _s(_col(row, 27)),
            "dd_references_count": _i(_col(row, 28)),
            "dd_is_investable": _b(_col(row, 29)),
            # GitHub signals
            "github_repo_count": _i(_col(row, 30)),
            "github_contributors": _i(_col(row, 31)),
            "github_commits_90d": _i(_col(row, 32)),
            "github_ci_present": _b(_col(row, 33)),
            # Research signals
            "huggingface_model_hits": _i(_col(row, 34)),
            "openalex_hits": _i(_col(row, 35)),
            "semantic_scholar_hits": _i(_col(row, 36)),
            "linkedin_hits": _i(_col(row, 37)),
            "researchgate_hits": _i(_col(row, 38)),
            # Pre-computed scores
            "outreach_now_score": _f(_col(row, 39)),
            "venture_upside_score": _f(_col(row, 40)),
            # Coverage
            "profile_coverage_score": _i(_col(row, 19)),
            "known_url_count": _i(_col(row, 43)),
        })
    return out


# Fields that come from the overview spreadsheet
_OVERVIEW_FIELDS = (
    "technology_domains", "market_domains", "categories",
    "member_count", "member_examples", "member_roles",
    "github_repo_count", "github_contributors", "github_commits_90d", "github_ci_present",
    "huggingface_model_hits", "openalex_hits", "semantic_scholar_hits",
    "dd_key_roles", "dd_references_count", "dd_is_investable",
    "outreach_now_score", "venture_upside_score",
    "profile_coverage_score", "known_url_count",
    "linkedin_hits", "researchgate_hits",
)


def _normalize_key(name: str, uni: str) -> str:
    return f"{name.strip().casefold()}|{uni.strip().casefold()}"


def _upsert(session: Session, data: dict, existing: dict[str, Initiative]) -> tuple[bool, Initiative]:
    """Insert new or update existing initiative. Returns (is_new, initiative)."""
    key = _normalize_key(data["name"], data["uni"])
    is_overview = data.get("sheet_source") == "overview"

    if key in existing:
        init = existing[key]
        # Update text fields if the new data has more info
        for field in ("sector", "mode", "description", "website", "email", "team_page",
                      "team_size", "linkedin", "github_org", "key_repos", "sponsors",
                      "competitions", "relevance"):
            new_val = data.get(field, "")
            old_val = getattr(init, field, "") or ""
            if new_val and (not old_val or (data["sheet_source"] == "spin_off_targets" and field != "relevance")):
                setattr(init, field, new_val)
        # Overview fields: always overwrite (they come from the pipeline)
        if is_overview:
            for field in _OVERVIEW_FIELDS:
                val = data.get(field)
                if val is not None:
                    setattr(init, field, val)
        # Merge extra links
        try:
            old_links = json.loads(init.extra_links_json or "{}")
        except (json.JSONDecodeError, TypeError):
            old_links = {}
        try:
            new_links = json.loads(data.get("extra_links_json", "{}"))
        except (json.JSONDecodeError, TypeError):
            new_links = {}
        merged = {**old_links, **{k: v for k, v in new_links.items() if v}}
        init.extra_links_json = json.dumps(merged)
        return False, init
    else:
        # For overview-only inserts, fill required Initiative fields with defaults
        if is_overview:
            data.setdefault("sector", "")
            data.setdefault("mode", "")
            data.setdefault("website", "")
            data.setdefault("email", "")
            data.setdefault("team_page", "")
            data.setdefault("team_size", "")
            data.setdefault("linkedin", "")
            data.setdefault("github_org", "")
            data.setdefault("key_repos", "")
            data.setdefault("sponsors", "")
            data.setdefault("competitions", "")
            data.setdefault("relevance", "")
        init = Initiative(**data)
        session.add(init)
        existing[key] = init
        return True, init


def import_xlsx(file_path: str | Path, session: Session) -> ImportResult:
    """Import both sheets from the enriched XLSX. Upserts by name+uni."""
    file_path = Path(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Load existing initiatives for dedup
    all_existing = session.execute(select(Initiative)).scalars().all()
    existing_map: dict[str, Initiative] = {
        _normalize_key(i.name, i.uni): i for i in all_existing
    }

    spin_off_rows: list[dict] = []
    all_init_rows: list[dict] = []
    overview_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lower = sheet_name.casefold()
        if "spin" in lower and "off" in lower:
            spin_off_rows = _parse_spin_off_sheet(ws)
        elif "all" in lower and "init" in lower:
            all_init_rows = _parse_all_initiatives_sheet(ws)
        elif lower == "initiatives":
            overview_rows = _parse_overview_sheet(ws)

    wb.close()

    new_count = 0
    updated_count = 0

    # Import spin-off targets first (higher quality data)
    for data in spin_off_rows:
        is_new, _ = _upsert(session, data, existing_map)
        if is_new:
            new_count += 1
        else:
            updated_count += 1

    # Then all initiatives (fills gaps, adds relevance rating)
    for data in all_init_rows:
        is_new, _ = _upsert(session, data, existing_map)
        if is_new:
            new_count += 1
        else:
            updated_count += 1

    # Then overview data (adds enriched signals to existing entries)
    for data in overview_rows:
        is_new, _ = _upsert(session, data, existing_map)
        if is_new:
            new_count += 1
        else:
            updated_count += 1

    session.commit()

    return ImportResult(
        total_imported=new_count + updated_count,
        spin_off_count=len(spin_off_rows),
        all_initiatives_count=len(all_init_rows) + len(overview_rows),
        duplicates_updated=updated_count,
    )
